import os
import json
import webbrowser
from kivy.app import App
from kivy.uix.label import Label
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.core.window import Window
from kivy.properties import StringProperty
from openpyxl import load_workbook
from xlcalculator import ModelCompiler, Evaluator


class DragDropLabel(Label):
    def __init__(self, callback, **kwargs):
        super().__init__(**kwargs)
        self.callback = callback
        Window.bind(on_dropfile=self._on_file_drop)
        self.text = "Drop .xlsx file here"

    def _on_file_drop(self, window, file_path):
        file_path = file_path.decode("utf-8")
        if file_path.endswith(".xlsx"):
            self.callback(file_path)
        else:
            self.text = "Invalid file type"


class ExcelFormulaExtractorApp(App):
    output_path = StringProperty("")

    def build(self):
        self.title = "Excel Formula Extractor"
        layout = BoxLayout(orientation="vertical", spacing=10, padding=10)

        self.status_label = Label(text="Drop a file to begin", size_hint=(1, 0.2))
        self.drag_label = DragDropLabel(self.extract_formulas, size_hint=(1, 0.5))
        self.open_button = Button(text="Open output folder", size_hint=(1, 0.2))
        self.open_button.bind(on_release=self.open_output_folder)

        layout.add_widget(self.drag_label)
        layout.add_widget(self.status_label)
        layout.add_widget(self.open_button)
        return layout

    def extract_formulas(self, file_path):
        try:
            workbook = load_workbook(filename=file_path, data_only=False)
            sheet_name = "Formulas"
            if sheet_name not in workbook.sheetnames:
                raise ValueError("No sheet named 'Formulas' found.")

            ws = workbook[sheet_name]
            formulas = []

            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and cell.value:
                        formulas.append({
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "formula": cell.value
                        })

            output_dir = os.path.dirname(file_path)
            extracted_path = os.path.join(output_dir, "extracted_formulas.json")
            with open(extracted_path, "w") as f:
                json.dump(formulas, f, indent=2)

            evaluated = self.evaluate_with_xlcalculator(file_path, formulas)
            evaluated_path = os.path.join(output_dir, "evaluated_formulas.json")
            with open(evaluated_path, "w") as f:
                json.dump(evaluated, f, indent=2)

            self.output_path = evaluated_path
            self.status_label.text = "Exported formulas and results successfully"
            self.drag_label.text = "Drop .xlsx file here"

        except Exception as e:
            self.status_label.text = f"Error: {str(e)}"

    def evaluate_with_xlcalculator(self, excel_path, formula_list):
        compiler = ModelCompiler()
        model = compiler.read_and_parse_archive(excel_path)
        evaluator = Evaluator(model)

        results = []
        for entry in formula_list:
            ref = f"{entry['sheet']}!{entry['cell']}"
            try:
                val = evaluator.evaluate(ref)
                value = float(val.value) if hasattr(val, "value") and isinstance(val.value, (int, float)) else str(val)
            except Exception as e:
                value = f"Error: {e}"

            results.append({
                "cell": ref,
                "formula": entry["formula"],
                "evaluated_result": value
            })
        return results

    def open_output_folder(self, *args):
        if self.output_path and os.path.exists(self.output_path):
            folder = os.path.dirname(self.output_path)
            webbrowser.open(f"file://{folder}")
        else:
            self.status_label.text = "No output file available."


if __name__ == "__main__":
    ExcelFormulaExtractorApp().run()
