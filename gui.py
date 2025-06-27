import os
import json
import webbrowser
from kivy.app import App
from kivy.uix.label import Label
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.popup import Popup
from kivy.core.window import Window
from kivy.uix.floatlayout import FloatLayout
from kivy.properties import StringProperty
from openpyxl import load_workbook


class DragDropLabel(Label):
    def __init__(self, callback, **kwargs):
        super().__init__(**kwargs)
        self.callback = callback
        Window.bind(on_dropfile=self._on_file_drop)
        self.text = "file drop here"

    def _on_file_drop(self, window, file_path):
        file_path = file_path.decode("utf-8")
        if file_path.endswith(".xlsx"):
            self.callback(file_path)
        else:
            self.text = "invalid file type"


class ExcelFormulaExtractorApp(App):
    output_path = StringProperty("")

    def build(self):
        self.title = "Excel Formula Extractor"
        layout = BoxLayout(orientation="vertical", spacing=10, padding=10)

        self.status_label = Label(text="Drop a file to begin", size_hint=(1, 0.2))

        self.drag_label = DragDropLabel(self.extract_formulas, size_hint=(1, 0.5))

        self.open_button = Button(text="Open output foler", size_hint=(1, 0.2))
        self.open_button.bind(on_release=self.open_output_folder)

        layout.add_widget(self.drag_label)
        layout.add_widget(self.status_label)
        layout.add_widget(self.open_button)

        return layout

    def extract_formulas(self, file_path):
        try:
            workbook = load_workbook(filename=file_path, data_only=False)
            formulas = []

            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f':
                            formulas.append({
                                "sheet": sheet,
                                "cell": cell.coordinate,
                                "formula": cell.value
                            })

            output_filename = "extracted_formulas.json"
            output_path = os.path.join(os.path.dirname(file_path), output_filename)
            with open(output_path, "w") as f:
                json.dump(formulas, f, indent=2)

            self.output_path = output_path
            self.status_label.text = f"Exported to: {output_filename}"
            self.drag_label.text = "file drop here"

        except Exception as e:
            self.status_label.text = f" Error: {str(e)}"

    def open_output_folder(self, *args):
        if self.output_path and os.path.exists(self.output_path):
            folder = os.path.dirname(self.output_path)
            webbrowser.open(f"file://{folder}")
        else:
            self.status_label.text = "No output file available."


if __name__ == "__main__":
    ExcelFormulaExtractorApp().run()
