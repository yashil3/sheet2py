import os
from openpyxl import load_workbook


def extract_data_and_formulas_from_excel(file_path):
    """Extracts formulas, cell data, key-value mappings, and named ranges from an Excel file.
    Detects explicit Key/Value tables (two columns labeled 'Key' and 'Value') and
    also implicit label/value pairs without headers (e.g., labels in one column and values in the next).
    Additionally detects criteria matrices with a 'Key' column and multiple header columns, mapping
    each data cell to a derived key of the form '<row_key>:<column_header>'.
    Also extracts named ranges defined in the workbook.
    Returns per sheet:
      - formulas: list of {cell, formula}
      - data: dict of cell_reference -> value
      - key_values: dict of key -> value
      - cell_to_key: dict of value_cell_reference -> key (e.g., 'B12' -> 'Engine' or 'Color:Choice1')
    And at workbook level:
      - named_ranges: dict of named_range_name -> (sheet_name, cell_reference)
    """
    workbook = load_workbook(filename=file_path, data_only=True)  # Use data_only=True to get values
    formulas_workbook = load_workbook(filename=file_path, data_only=False)

    # Extract named ranges from the workbook
    named_ranges = {}
    try:
        # The correct way to access defined names in openpyxl
        if hasattr(workbook, 'defined_names') and workbook.defined_names:
            for name, named_range in workbook.defined_names.items():
                # Get the target of the named range
                target = named_range.attr_text
                if target:
                    # Parse the target (e.g., "Sheet1!$A$1" or "Sheet1!$A$1:$B$10")
                    if '!' in target:
                        sheet_name, cell_ref = target.split('!', 1)
                        # Remove quotes from sheet name if present
                        if sheet_name.startswith("'") and sheet_name.endswith("'"):
                            sheet_name = sheet_name[1:-1]
                        # Remove absolute markers from cell reference
                        cell_ref = cell_ref.replace('$', '')
                        named_ranges[name] = (sheet_name, cell_ref)
    except Exception as e:
        # If named ranges can't be extracted, continue without them
        print(f"Warning: Could not extract named ranges: {e}")

    extracted_data = {}

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        formulas_ws = formulas_workbook[sheet_name]

        formulas_in_sheet = []
        data_in_sheet = {}
        key_values = {}
        cell_to_key = {}

        # Collect raw cell values
        for row in ws.iter_rows():
            for cell in row:
                data_in_sheet[cell.coordinate] = cell.value

        # Collect formulas (use non-data_only workbook)
        for row in formulas_ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value:
                    formulas_in_sheet.append({
                        "cell": cell.coordinate,
                        "formula": f"={cell.value}" if not str(cell.value).startswith('=') else str(cell.value)
                    })

        # Try to detect a Key/Value header row within the first few rows
        header_found = False
        header_row_idx = None
        key_col_letter = None
        value_col_letter = None

        max_scan_rows = min(10, ws.max_row)
        for r in range(1, max_scan_rows + 1):
            row_values = {cell.column_letter: (cell.value.strip() if isinstance(cell.value, str) else cell.value)
                          for cell in ws[r] if cell.value is not None}
            key_cols = [col for col, val in row_values.items() if isinstance(val, str) and val.lower() == 'key']
            value_cols = [col for col, val in row_values.items() if isinstance(val, str) and val.lower() == 'value']
            if key_cols and value_cols:
                key_col_letter = key_cols[0]
                value_col_letter = value_cols[0]
                header_row_idx = r
                header_found = True
                break

        # If key/value layout found, build the mappings
        if header_found and key_col_letter and value_col_letter and header_row_idx:
            for r in range(header_row_idx + 1, ws.max_row + 1):
                key_cell = f"{key_col_letter}{r}"
                val_cell = f"{value_col_letter}{r}"
                key = ws[key_cell].value
                val = ws[val_cell].value
                if key is None:
                    continue
                if isinstance(key, str):
                    key = key.strip()
                if key == "":
                    continue
                key_values[key] = val
                cell_to_key[val_cell] = key
        else:
            # Heuristic: find adjacent label/value columns without headers
            # Strategy: scan pairs of adjacent columns and count rows where left cell is a non-empty string
            # and right cell is non-empty (numeric or string). Pick the pair with max matches.
            max_matches = 0
            best_pair = None  # (label_col_letter, value_col_letter)
            # Limit scan to first 5 non-empty columns to keep this cheap
            non_empty_cols = list(range(1, min(ws.max_column, 10) + 1))
            for col_idx in non_empty_cols:
                if col_idx >= ws.max_column:
                    continue
                label_col = col_idx
                value_col = col_idx + 1
                matches = 0
                for r in range(1, ws.max_row + 1):
                    lcell = ws.cell(row=r, column=label_col)
                    vcell = ws.cell(row=r, column=value_col)
                    lval = lcell.value
                    vval = vcell.value
                    if isinstance(lval, str) and lval.strip() not in ("Key", "Value") and vval is not None:
                        matches += 1
                if matches > max_matches and matches >= 3:  # need at least a few rows to be credible
                    max_matches = matches
                    best_pair = (label_col, value_col)
            if best_pair:
                label_col, value_col = best_pair
                for r in range(1, ws.max_row + 1):
                    lcell = ws.cell(row=r, column=label_col)
                    vcell = ws.cell(row=r, column=value_col)
                    key = lcell.value
                    val = vcell.value
                    if isinstance(key, str):
                        k = key.strip()
                        if k:
                            key_values[k] = val
                            cell_to_key[vcell.coordinate] = k

        # Detect criteria matrices: find a header row containing a 'Key' header and other non-empty headers.
        # For each row under it, map value cells to derived semantic keys '<row_key>:<header>'.
        criteria_header_row = None
        criteria_key_col_idx = None
        criteria_headers = {}  # col_index -> header string

        for r in range(1, min(ws.max_row, 20) + 1):
            # Build a map of column index -> header value for row r
            headers_this_row = {}
            key_col_idx_candidate = None
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str):
                    vv = v.strip()
                    headers_this_row[c] = vv
                    if vv.lower() == 'key':
                        key_col_idx_candidate = c
            if key_col_idx_candidate is not None:
                # Build headers excluding the 'Key' column (and exclude a 'Value' column if present on same row)
                local_headers = {}
                for c, hv in headers_this_row.items():
                    if c == key_col_idx_candidate:
                        continue
                    if hv.lower() == 'value':
                        # skip VALUE column to avoid duplicating simple key/value pairs
                        continue
                    if hv:
                        local_headers[c] = hv
                if local_headers:
                    criteria_header_row = r
                    criteria_key_col_idx = key_col_idx_candidate
                    criteria_headers = local_headers
                    break

        if criteria_header_row and criteria_key_col_idx and criteria_headers:
            # Iterate rows under header until a blank key
            for rr in range(criteria_header_row + 1, ws.max_row + 1):
                row_key_val = ws.cell(row=rr, column=criteria_key_col_idx).value
                if not isinstance(row_key_val, str) or not row_key_val.strip():
                    continue
                row_key = row_key_val.strip()
                for c_idx, header_text in criteria_headers.items():
                    vcell = ws.cell(row=rr, column=c_idx)
                    coord = vcell.coordinate
                    derived_key = f"{row_key}:{header_text}"
                    cell_to_key[coord] = derived_key
                    key_values[derived_key] = vcell.value

        extracted_data[sheet_name] = {
            "formulas": formulas_in_sheet,
            "data": data_in_sheet,
            "key_values": key_values,
            "cell_to_key": cell_to_key,
        }

    # Add named ranges at the workbook level
    extracted_data['_named_ranges'] = named_ranges

    return extracted_data