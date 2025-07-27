import os
import json
import argparse
from pathlib import Path
from openpyxl import load_workbook
from typing import Dict, List, Any, Tuple, Optional

def find_key_value_columns(ws) -> Optional[Tuple[int, int]]:
    """Find columns containing 'key' and 'value' headers."""
    key_col = None
    value_col = None
    
    # Search through the worksheet to find key/value headers
    for row in ws.iter_rows(max_row=10):  # Search first 10 rows
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                cell_value = cell.value.lower().strip()
                if cell_value in ['key', 'keys']:
                    key_col = cell.column
                elif cell_value in ['value', 'values']:
                    value_col = cell.column
    
    return (key_col, value_col) if key_col and value_col else None

def extract_key_value_pairs(ws, key_col: int, value_col: int) -> Dict[str, Any]:
    """Extract key-value pairs from specified columns."""
    key_value_pairs = {}
    
    # Start from row after headers
    for row in ws.iter_rows(min_row=2):
        key_cell = row[key_col - 1] if len(row) >= key_col else None
        value_cell = row[value_col - 1] if len(row) >= value_col else None
        
        if key_cell and key_cell.value is not None:
            key = str(key_cell.value).strip()
            value = value_cell.value if value_cell else None
            
            if key:  # Only add non-empty keys
                key_value_pairs[key] = value
    
    return key_value_pairs

def find_customer_preference_data(ws) -> Optional[Dict[str, int]]:
    """Find the actual customer preference values (not formulas/weights)."""
    columns = {}
    
    # Look for "Customer preferences" header and the data beneath it
    for row_idx, row in enumerate(ws.iter_rows(max_row=15), 1):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                cell_value = cell.value.lower().strip()
                if 'customer preferences' in cell_value:
                    # Found the header, now look for the data structure below
                    columns['preference'] = cell.column
                    columns['header_row'] = row_idx
                    
                    # Look at the next few rows to find where the actual data starts
                    # and verify it contains text values (not just numbers)
                    for data_row_idx in range(row_idx + 1, min(row_idx + 10, ws.max_row + 1)):
                        data_row = list(ws.iter_rows(min_row=data_row_idx, max_row=data_row_idx))[0]
                        if len(data_row) > cell.column - 1:
                            first_cell = data_row[cell.column - 1]
                            if first_cell.value and isinstance(first_cell.value, str):
                                # Check if the adjacent cells contain preference values (text, not numbers)
                                has_text_values = False
                                for col_offset in range(1, 6):  # Check next 5 columns
                                    if len(data_row) >= cell.column + col_offset:
                                        value_cell = data_row[cell.column + col_offset - 1]
                                        if value_cell and value_cell.value is not None:
                                            value = str(value_cell.value).strip()
                                            # Check if it's a text value (not a number, not a formula)
                                            if (value and not value.startswith('=') and 
                                                not value.isdigit() and 
                                                value.upper() not in ['N/A', 'NULL', '']):
                                                has_text_values = True
                                                break
                                
                                if has_text_values:
                                    columns['data_start_row'] = data_row_idx
                                    break
                    break
        if 'preference' in columns and 'data_start_row' in columns:
            break
    
    return columns if 'preference' in columns and 'data_start_row' in columns else None

def find_preference_weights_data(ws) -> Optional[Dict[str, int]]:
    """Find customer preferences with weights and calculations."""
    columns = {}
    
    # Search for weight-related headers or numeric data
    for row in ws.iter_rows(max_row=10):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                cell_value = cell.value.lower().strip()
                if any(keyword in cell_value for keyword in ['customer preferences', 'preference', 'criteria']):
                    columns['preference'] = cell.column
                elif 'weight' in cell_value:
                    columns['weight'] = cell.column
                elif 'relative' in cell_value and 'weight' in cell_value:
                    columns['relative_weight'] = cell.column
    
    # Verify this looks like a weights sheet (has numeric data)
    if 'preference' in columns:
        # Check if there are numeric values in the data
        has_numeric_data = False
        for row in ws.iter_rows(min_row=2, max_row=10):
            if len(row) >= columns['preference']:
                pref_cell = row[columns['preference'] - 1]
                if pref_cell and pref_cell.value:
                    # Look for numeric values in adjacent columns
                    for col_offset in range(1, 5):
                        if len(row) >= columns['preference'] + col_offset:
                            value_cell = row[columns['preference'] + col_offset - 1]
                            if value_cell and value_cell.value is not None:
                                try:
                                    float(value_cell.value)
                                    has_numeric_data = True
                                    break
                                except (ValueError, TypeError):
                                    continue
                if has_numeric_data:
                    break
        
        if not has_numeric_data:
            return None
    
    return columns if 'preference' in columns and ('weight' in columns or has_numeric_data) else None

def extract_customer_preference_values(ws, columns: Dict[str, int]) -> Dict[str, Any]:
    """Extract the actual customer preference values (like 'GER', 'Orange', etc.)."""
    preferences = {}
    
    # Start from the data start row
    start_row = columns['data_start_row']
    pref_col = columns['preference']
    
    for row in ws.iter_rows(min_row=start_row):
        pref_cell = row[pref_col - 1] if len(row) >= pref_col else None
        
        if pref_cell and pref_cell.value is not None:
            preference_name = str(pref_cell.value).strip()
            
            # Skip empty preferences or headers
            if not preference_name or preference_name.lower() in ['customer preferences', 'preference', 'weight']:
                continue
                
            # Extract values from the same row (next columns)
            pref_values = []
            for col_offset in range(1, 6):  # Check next 5 columns
                if len(row) >= pref_col + col_offset:
                    value_cell = row[pref_col + col_offset - 1]
                    if value_cell and value_cell.value is not None:
                        value = str(value_cell.value).strip()
                        # Only add non-empty, non-formula values
                        if value and not value.startswith('=') and value.upper() not in ['N/A', 'NULL', '']:
                            pref_values.append(value)
            
            if pref_values:  # Only add if we found actual values
                preferences[preference_name] = {
                    "preference_values": pref_values
                }
    
    return {
        "preferences": preferences,
        "type": "values_only"
    }

def extract_preferences_with_weights(ws, columns: Dict[str, int]) -> Dict[str, Any]:
    """Extract customer preferences with weights and calculations."""
    preferences = {}
    total_weight = 0
    
    # First pass: collect all data and calculate total weight
    for row in ws.iter_rows(min_row=2):
        pref_cell = row[columns['preference'] - 1] if len(row) >= columns['preference'] else None
        
        if pref_cell and pref_cell.value is not None:
            preference = str(pref_cell.value).strip()
            
            # Extract weight
            weight = 0
            if 'weight' in columns:
                weight_cell = row[columns['weight'] - 1] if len(row) >= columns['weight'] else None
                try:
                    weight = int(weight_cell.value) if weight_cell and weight_cell.value is not None else 0
                except (ValueError, TypeError):
                    weight = 0
            
            if preference and weight > 0:
                preferences[preference] = {
                    "weight": weight
                }
                total_weight += weight
    
    # Second pass: calculate relative weights
    if total_weight > 0:
        for preference in preferences:
            relative_weight = preferences[preference]["weight"] / total_weight
            preferences[preference]["relative_weight"] = round(relative_weight, 4)
    
    return {
        "preferences": preferences,
        "total_weight": total_weight,
        "type": "weights_and_calculations"
    }

def merge_preference_data(values_data: Dict[str, Any], weights_data: Dict[str, Any]) -> Dict[str, Any]:
    """Merge customer preference values with their weights."""
    if not values_data.get("preferences") or not weights_data.get("preferences"):
        return values_data if values_data.get("preferences") else weights_data
    
    merged_preferences = {}
    
    # Start with weights data as base
    for pref_name, weight_info in weights_data["preferences"].items():
        merged_preferences[pref_name] = weight_info.copy()
        
        # Add preference values if they exist
        if pref_name in values_data["preferences"]:
            merged_preferences[pref_name]["preference_values"] = values_data["preferences"][pref_name]["preference_values"]
        else:
            merged_preferences[pref_name]["preference_values"] = []
    
    # Add any preference values that don't have weights
    for pref_name, value_info in values_data["preferences"].items():
        if pref_name not in merged_preferences:
            merged_preferences[pref_name] = {
                "weight": 0,
                "relative_weight": 0.0,
                "preference_values": value_info["preference_values"]
            }
    
    return {
        "preferences": merged_preferences,
        "total_weight": weights_data.get("total_weight", 0)
    }

def find_customer_preference_table(ws) -> Optional[Dict[str, int]]:
    """Find a table of customer preferences with format 'preference | value1 | value2 | etc'."""
    columns = {}
    
    # Look for "Customer preferences" header row
    for row_idx, row in enumerate(ws.iter_rows(max_row=20), 1):
        for cell_idx, cell in enumerate(row, 1):
            if cell.value and isinstance(cell.value, str):
                cell_value = cell.value.lower().strip()
                if 'customer preferences' in cell_value:
                    columns['header_row'] = row_idx
                    columns['start_col'] = cell_idx
                    
                    # Found the header, now check if the next row contains preference names
                    if row_idx + 1 <= ws.max_row:
                        # Get first item in the next row (under "Customer preferences")
                        next_row = list(ws.iter_rows(min_row=row_idx + 1, max_row=row_idx + 1))[0]
                        if len(next_row) >= cell_idx:
                            first_item = next_row[cell_idx - 1]
                            if first_item and first_item.value:
                                columns['data_start_row'] = row_idx + 1
                                return columns
    
    return None

def extract_customer_preference_table(ws, columns: Dict[str, int]) -> Dict[str, Any]:
    """Extract customer preferences from a table format."""
    preferences = {}
    
    start_row = columns['data_start_row']
    start_col = columns['start_col']
    
    # Determine how many rows to process (stop at blank row)
    end_row = start_row
    for row_idx in range(start_row, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx))[0]
        if len(row) >= start_col and row[start_col - 1].value:
            end_row = row_idx
        else:
            break
    
    # Process each row in the preference table
    for row_idx in range(start_row, end_row + 1):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx))[0]
        if len(row) < start_col:
            continue
            
        pref_name_cell = row[start_col - 1]
        if not pref_name_cell.value:
            continue
            
        pref_name = str(pref_name_cell.value).strip()
        
        # Extract values from cells to the right of preference name
        pref_values = []
        col_idx = start_col + 1  # Start from the next column
        while col_idx <= len(row):
            value_cell = row[col_idx - 1]
            if value_cell.value and str(value_cell.value).strip():
                pref_values.append(str(value_cell.value).strip())
            col_idx += 1
            
        # Only add preferences with actual values
        if pref_name and pref_values:
            preferences[pref_name] = {
                "preference_values": pref_values
            }
    
    return {
        "preferences": preferences,
        "type": "values_only"
    }

def extract_formulas_and_data(file_path: str) -> Dict[str, Any]:
    """Extract formulas and key-value data from a single Excel file."""
    workbook = load_workbook(filename=file_path, data_only=False)
    result = {
        "formulas": [],
        "key_value_data": {},
        "customer_preferences": {}
    }

    customer_values_data = None
    customer_weights_data = None
    
    print(f"Processing {len(workbook.sheetnames)} sheets: {workbook.sheetnames}")

    # First pass: Check for dedicated customer preference table
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        
        # Try to find a dedicated customer preference table 
        # (like the one in your screenshot)
        pref_table = find_customer_preference_table(ws)
        if pref_table:
            print(f"  Found customer preference TABLE in {sheet_name}")
            customer_values_data = extract_customer_preference_table(ws, pref_table)
            if customer_values_data["preferences"]:
                print(f"  Extracted {len(customer_values_data['preferences'])} preference values from table")
                break  # Found what we need, stop looking further
    
    # Second pass: Process everything else
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        print(f"  Analyzing sheet: {sheet_name}")
        
        # Extract formulas
        sheet_formulas = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':
                    sheet_formulas.append({
                        "sheet": sheet_name,
                        "cell": cell.coordinate,
                        "formula": cell.value
                    })
        
        result["formulas"].extend(sheet_formulas)
        print(f"    Found {len(sheet_formulas)} formulas")
        
        # Extract key-value pairs
        kv_columns = find_key_value_columns(ws)
        if kv_columns:
            key_col, value_col = kv_columns
            key_value_pairs = extract_key_value_pairs(ws, key_col, value_col)
            if key_value_pairs:
                result["key_value_data"][sheet_name] = key_value_pairs
                print(f"    Found {len(key_value_pairs)} key-value pairs")
        
        # Only look for weights if we haven't already found a preference table
        if not customer_values_data:
            # Extract customer preference values (actual preferences like "GER", "Orange")
            pref_values_columns = find_customer_preference_data(ws)
            if pref_values_columns:
                print(f"    Found customer preference VALUES in {sheet_name}")
                customer_values_data = extract_customer_preference_values(ws, pref_values_columns)
                if customer_values_data["preferences"]:
                    print(f"    Extracted {len(customer_values_data['preferences'])} preference values")
        
        # Extract customer preference weights (weights and calculations)
        pref_weights_columns = find_preference_weights_data(ws)
        if pref_weights_columns:
            print(f"    Found customer preference WEIGHTS in {sheet_name}")
            customer_weights_data = extract_preferences_with_weights(ws, pref_weights_columns)
            if customer_weights_data["preferences"]:
                print(f"    Extracted {len(customer_weights_data['preferences'])} preference weights")
    
    # Merge customer preference values with weights
    if customer_values_data or customer_weights_data:
        print("  Merging preference data...")
        merged_data = merge_preference_data(
            customer_values_data or {"preferences": {}},
            customer_weights_data or {"preferences": {}}
        )
        if merged_data["preferences"]:
            result["customer_preferences"]["Combined"] = merged_data
            print(f"  Final merged preferences: {len(merged_data['preferences'])}")
    
    return result

def batch_process_files(input_dir: str, output_file: str) -> None:
    """Process all Excel files in the input directory."""
    input_path = Path(input_dir)
    all_data: Dict[str, Dict[str, Any]] = {}
    
    # Find all Excel files
    excel_files = list(input_path.glob("**/*.xlsx"))
    
    if not excel_files:
        print(f"No Excel files found in {input_dir}")
        return

    # Process each file
    for file_path in excel_files:
        print(f"Processing: {file_path.name}")
        try:
            file_data = extract_formulas_and_data(str(file_path))
            all_data[file_path.name] = file_data
            
            formula_count = len(file_data["formulas"])
            kv_count = sum(len(sheet_data) for sheet_data in file_data["key_value_data"].values())
            pref_count = sum(len(sheet_data["preferences"]) for sheet_data in file_data["customer_preferences"].values())
            
            print(f"Found {formula_count} formulas, {kv_count} key-value pairs, and {pref_count} customer preferences")
            
        except Exception as e:
            print(f"Error processing {file_path.name}: {str(e)}")
            continue

    # Save results
    with open(output_file, 'w') as f:
        json.dump(all_data, f, indent=2)
    
    print(f"\nProcessing complete!")
    print(f"Processed {len(excel_files)} files")
    print(f"Results saved to: {output_file}")

def main():
    parser = argparse.ArgumentParser(description='Batch Excel Formula and Data Extractor')
    parser.add_argument('input_dir', help='Directory containing Excel files')
    parser.add_argument('--output', '-o', 
                       default='extracted_data.json',
                       help='Output JSON file path (default: extracted_data.json)')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.input_dir):
        print(f"Error: Input directory '{args.input_dir}' does not exist")
        return

    batch_process_files(args.input_dir, args.output)

if __name__ == "__main__":
    main()