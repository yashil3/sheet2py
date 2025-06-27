import os
import json
import argparse
from pathlib import Path
from openpyxl import load_workbook
from typing import Dict, List, Any, Tuple, Optional

def find_key_value_columns(ws) -> Optional[Tuple[int, int]]:
    """Find columns containing 'key' and 'value' headers."""
    key_col = None
    value_col = None
    
    # Search through the worksheet to find key/value headers
    for row in ws.iter_rows(max_row=10):  # Search first 10 rows
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                cell_value = cell.value.lower().strip()
                if cell_value in ['key', 'keys']:
                    key_col = cell.column
                elif cell_value in ['value', 'values']:
                    value_col = cell.column
    
    return (key_col, value_col) if key_col and value_col else None

def extract_key_value_pairs(ws, key_col: int, value_col: int) -> Dict[str, Any]:
    """Extract key-value pairs from specified columns."""
    key_value_pairs = {}
    
    # Start from row after headers
    for row in ws.iter_rows(min_row=2):
        key_cell = row[key_col - 1] if len(row) >= key_col else None
        value_cell = row[value_col - 1] if len(row) >= value_col else None
        
        if key_cell and key_cell.value is not None:
            key = str(key_cell.value).strip()
            value = value_cell.value if value_cell else None
            
            if key:  # Only add non-empty keys
                key_value_pairs[key] = value
    
    return key_value_pairs

def find_preference_weight_columns(ws) -> Optional[Tuple[int, int]]:
    """Find columns containing 'preference' and 'weight' headers."""
    pref_col = None
    weight_col = None
    
    # Search through the worksheet to find preference/weight headers
    for row in ws.iter_rows(max_row=10):  # Search first 10 rows
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                cell_value = cell.value.lower().strip()
                if 'customer preferences' in cell_value:
                    pref_col = cell.column
                elif 'weight' in cell_value:
                    weight_col = cell.column
    
    return (pref_col, weight_col) if pref_col and weight_col else None

def extract_preferences_with_weights(ws, pref_col: int, weight_col: int) -> Dict[str, Any]:
    """Extract customer preferences with weights and calculate relative weights."""
    preferences = {}
    total_weight = 0
    
    # First pass: collect all data and calculate total weight
    for row in ws.iter_rows(min_row=2):
        pref_cell = row[pref_col - 1] if len(row) >= pref_col else None
        weight_cell = row[weight_col - 1] if len(row) >= weight_col else None
        
        if pref_cell and pref_cell.value is not None:
            preference = str(pref_cell.value).strip()
            
            try:
                weight = int(weight_cell.value) if weight_cell and weight_cell.value is not None else 0
            except (ValueError, TypeError):
                weight = 0
                
            if preference and weight > 0:  # Only add valid preferences with positive weights
                preferences[preference] = {"weight": weight}
                total_weight += weight
    
    # Second pass: calculate relative weights
    if total_weight > 0:
        for preference in preferences:
            relative_weight = preferences[preference]["weight"] / total_weight
            preferences[preference]["relative_weight"] = round(relative_weight, 4)
    
    return {
        "preferences": preferences,
        "total_weight": total_weight
    }

def extract_formulas_and_data(file_path: str) -> Dict[str, Any]:
    """Extract formulas and key-value data from a single Excel file."""
    workbook = load_workbook(filename=file_path, data_only=False)
    result = {
        "formulas": [],
        "key_value_data": {},
        "customer_preferences": {}
    }

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        
        # Extract formulas
        sheet_formulas = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':
                    sheet_formulas.append({
                        "sheet": sheet_name,
                        "cell": cell.coordinate,
                        "formula": cell.value
                    })
        
        result["formulas"].extend(sheet_formulas)
        
        # Extract key-value pairs
        kv_columns = find_key_value_columns(ws)
        if kv_columns:
            key_col, value_col = kv_columns
            key_value_pairs = extract_key_value_pairs(ws, key_col, value_col)
            if key_value_pairs:
                result["key_value_data"][sheet_name] = key_value_pairs
        
        # Extract customer preferences with weights
        pref_columns = find_preference_weight_columns(ws)
        if pref_columns:
            pref_col, weight_col = pref_columns
            preference_data = extract_preferences_with_weights(ws, pref_col, weight_col)
            if preference_data["preferences"]:
                result["customer_preferences"][sheet_name] = preference_data
    
    return result

def batch_process_files(input_dir: str, output_file: str) -> None:
    """Process all Excel files in the input directory."""
    input_path = Path(input_dir)
    all_data: Dict[str, Dict[str, Any]] = {}
    
    # Find all Excel files
    excel_files = list(input_path.glob("**/*.xlsx"))
    
    if not excel_files:
        print(f"No Excel files found in {input_dir}")
        return

    # Process each file
    for file_path in excel_files:
        print(f"Processing: {file_path.name}")
        try:
            file_data = extract_formulas_and_data(str(file_path))
            all_data[file_path.name] = file_data
            
            formula_count = len(file_data["formulas"])
            kv_count = sum(len(sheet_data) for sheet_data in file_data["key_value_data"].values())
            pref_count = sum(len(sheet_data["preferences"]) for sheet_data in file_data["customer_preferences"].values())
            
            print(f"Found {formula_count} formulas, {kv_count} key-value pairs, and {pref_count} customer preferences")
            
        except Exception as e:
            print(f"Error processing {file_path.name}: {str(e)}")
            continue

    # Save results
    with open(output_file, 'w') as f:
        json.dump(all_data, f, indent=2)
    
    print(f"\nProcessing complete!")
    print(f"Processed {len(excel_files)} files")
    print(f"Results saved to: {output_file}")

def main():
    parser = argparse.ArgumentParser(description='Batch Excel Formula and Data Extractor')
    parser.add_argument('input_dir', help='Directory containing Excel files')
    parser.add_argument('--output', '-o', 
                       default='extracted_data.json',
                       help='Output JSON file path (default: extracted_data.json)')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.input_dir):
        print(f"Error: Input directory '{args.input_dir}' does not exist")
        return

    batch_process_files(args.input_dir, args.output)

if __name__ == "__main__":
    main()